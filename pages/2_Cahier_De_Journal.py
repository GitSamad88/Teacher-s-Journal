# !pip3 install translate
# !pip3 install babel
# !pip3 install googletrans
# !pip3 install py-translate
# !pip3 install deepl
# !pip3 install tabula-py
# !pip3 install openpyxl

import streamlit as st
import pandas as pd
import numpy as np

from datetime import datetime
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from babel.dates import format_datetime
from io import BytesIO
import warnings
import json
warnings.filterwarnings("ignore")


def fr_manuel_level(manuel, level):
    # mes_app1 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\Repa  fr dire faire 1aep final.csv")  # to be changed!
    mes_app1 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1VDIV3k4GraaUILZw46e9fDWuY4EGPeyCyRJJZ4ltyWs/gviz/tq?tqx=out:csv&sheet=Repa_fr_dire_faire_1aep_final")

    # mes_app2 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\repa fr mes app 2aep (Edited _6).csv")
    mes_app2 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1cXdwSlyjEp-_31qj4DBj__Jj6xfVE1QkJTwFI7b4a18/gviz/tq?tqx=out:csv&sheet=repa_fr_mes_app_2aep_(Edited_6)")

    # mes_app3 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\repartition_annuelle_fr_mes_apprentissage_3aep_final.csv")
    mes_app3 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1FEEDFVHEqVkoKA8ify7SIxWqxyRWn1Eu7hA89zk0mAs/gviz/tq?tqx=out:csv&sheet=repartition_annuelle_fr_mes_apprentissage_3aep_final")

    # mes_app4 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\repartiton_FR_Mes apprentissage_4AEP.csv")

    mes_app4 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1hAzdEag4ldicksVnWwLtxykKjZGHveInt2wjdc7AROU/gviz/tq?tqx=out:csv&sheet=repartiton_FR_Mes_apprentissage_4AEP")
    mes_app4["Dictèe"] = mes_app4["Orth / Dictée"]
    mes_app4.rename(columns={"Orth / Dictée": "Orthographe"}, inplace=True)
    # mes_app5 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\annuelle repa mes app 5aep (Edited _7).csv")
    mes_app5 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1JjobDWfXmLZK3QZHWFxbIINL02VtU2HWlZ5mPJg7y7U/gviz/tq?tqx=out:csv&sheet=annuelle_repa_mes_app_5aep_(Edited_7)")
    # mes_app6 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\fr repa\Mes apprentissage\Repa Fr Mes app 6 aep (Edited _3).csv")
    mes_app6 = pd.read_csv(
        "https://docs.google.com/spreadsheets/d/1B_YfR2CiXkGgipQd4GDnPVClfJtjHU5C99WJJmhSUAM/gviz/tq?tqx=out:csv&sheet=Repa_Fr_Mes_app_6_aep_(Edited_3)")
    Mes_app_manuel = [mes_app1, mes_app2, mes_app3, mes_app4, mes_app5, mes_app6]
    fr_manuels = {"Mes apprentissages": Mes_app_manuel, "L'oasis des mots": [], "Pour communiquer": [],
                  "L'école des mots": [], "Le nouvel espace": [], "Mon livre de français": []}

    try:
        return fr_manuels[str(manuel)][int(level) - 1]
    except:
        print("Ce manuel n'est pas disponible pour ce niveau, essayez un autre manuel!")
        print("Remarque: le niveau doit être entre 1 et 6!")


# Maths

def maths_manuel_level(manuel, level):
    moufid1 = pd.DataFrame()
    #moufid1 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Almoufid\Repa Maths Moufid 1aep.csv")
    moufid2 = pd.DataFrame()
    moufid3 = pd.DataFrame()
    moufid4 = pd.DataFrame()
    moufid5 = pd.DataFrame()
    #moufid5 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Almoufid\Repa Maths Almoufid 5aep .csv")
    moufid6 = pd.DataFrame()
    moufid_manuel = [moufid1, moufid2, moufid3, moufid4, moufid5, moufid6]

    # Jayed
    jayed1 = pd.DataFrame()
    jayed2 = pd.DataFrame()
    #jayed2 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Jayed\Repa Maths Jayed 2 aep.csv")
    jayed3 = pd.DataFrame()
    #jayed4 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Jayed\Repartition_Maths_Jayed_4aep.csv")
    jayed4 = pd.read_csv("https://docs.google.com/spreadsheets/d/1ckXxHHh5mCB_YMD11ObGEep0UfGJGPqpXDNY62_f0mc/gviz/tq?tqx=out:csv&sheet=Repartition_Maths_Jayed_4aep")
    jayed5 = pd.DataFrame()
    jayed6 = pd.DataFrame()
    #jayed6 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Jayed\Repa Maths Jayed 6 aep.csv")
    jayed_manuel = [jayed1, jayed2, jayed3, jayed4, jayed5, jayed6]

    # Fadaa:
    fada1 = pd.DataFrame()
    fada2 = pd.DataFrame()
    fada3 = pd.read_csv("https://docs.google.com/spreadsheets/d/16Tpk0FjPLs6jpbAjVa4DWC7L4VYxfruldvujbpWoJsc/gviz/tq?tqx=out:csv&sheet=Repa__3aep_Maths_Fada2")
    #fada3 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\Maths repa\Fada2\Repa__3aep_Maths_Fada2.csv")

    fada4 = pd.DataFrame()
    fada5 = pd.DataFrame()
    fada6 = pd.DataFrame()
    fadaa_manuel = [fada1, fada2, fada3, fada4, fada5, fada6]

    Maths_manuels = {"المفيد": moufid_manuel, "الجيد": jayed_manuel, "الفضاء": fadaa_manuel}

    try:
        if Maths_manuels[str(manuel)][int(level) - 1].empty:
            return "Ce manuel n'est pas disponible pour ce niveau, choisissez un autre manuel!"
        else:

            return Maths_manuels[str(manuel)][int(level) - 1]
    except:
        print("Ce manuel n'est pas disponible pour ce niveau, choisissez un autre manuel!")
        print("Remarque: le niveau doit être entre 1 et 6!")


# In[94]:


maths_manuel_level(manuel="الجيد", level=1)


# ### Eveil Scientifque:
def EvSc_manuel_level(level, manuel):
    # Jadid
    es_Jadid1 = pd.DataFrame()
    #es_Jadid1 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Jadid\Repa EvSc Jadid 1aep.csv")
    #es_Jadid1 = es_Jadid1[["U", "S", "Eveil Scientifique", "Sèance"]].dropna(axis=0)
    #es_Jadid1["Sèance"] = [json.loads(es_Jadid1["Sèance"][i]) for i in range(len(es_Jadid1["Sèance"]))]

    es_Jadid2 = pd.DataFrame()
    es_Jadid3 = pd.DataFrame()
    es_Jadid4 = pd.DataFrame()
    es_Jadid5 = pd.DataFrame()
    es_Jadid6 = pd.DataFrame()
    es_jadid_manuel = [es_Jadid1,
                       es_Jadid2,
                       es_Jadid3,
                       es_Jadid4,
                       es_Jadid5,
                       es_Jadid6]

    # Fadaa
    es_fadaa1 = pd.DataFrame()
    es_fadaa2 = pd.DataFrame()
    es_fadaa3 = pd.DataFrame()
    es_fadaa4 = pd.DataFrame()
    es_fadaa5 = pd.DataFrame()
    es:fadaa6 = pd.DataFrame()
    #es_fadaa6 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Fada2\Repa EvSc Fada2 6aep (Edited _3).csv")
    #es_fadaa6 = es_fadaa6[["U", "S", "Eveil Scientifique", "Sèance"]].dropna(axis=0)
    #es_fadaa6["Sèance"] = [json.loads(es_fadaa6["Sèance"][i]) for i in range(len(es_fadaa6["Sèance"]))]
    es_fadaa_manuel = [es_fadaa1,
                       es_fadaa2,
                       es_fadaa3,
                       es_fadaa4,
                       es_fadaa5,
                       es_fadaa6]

    # Manhal

    manhal1 = pd.DataFrame()
    manhal2 = pd.DataFrame()
    #manhal3 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Manhal\Repa EvSc Manhal 3aep (Edited _4).csv")
    manhal3=pd.read_csv("https://docs.google.com/spreadsheets/d/1Z37TLj6yc4HsVVZ9MY71aS6yqV2rGdpxkC1_RyZpoXY/"
                        "gviz/tq?tqx=out:csv&sheet=Repa_EvSc_Manhal_3aep_(Edited_4)")
    manhal3 = manhal3[["U", "S", "Eveil Scientifique", "Séance"]].dropna(axis=0)
    manhal3.rename(columns={"Séance": "Sèance"}, inplace=True)
    manhal3["Sèance"] = [json.loads(manhal3["Sèance"][i]) for i in range(len(manhal3["Sèance"]))]

    manhal4 = pd.DataFrame()
    manhal5 = pd.DataFrame()
    manhal6 = pd.DataFrame()

    manhal_manuel = [manhal1,
                     manhal2,
                     manhal3,
                     manhal4,
                     manhal5,
                     manhal6]
    # for df in manhal_manuel:
    #   df=df[["U","S","Eveil Scientifique","Sèance"]].dropna(axis=0)

    # Moufid

    es_moufid1 = pd.DataFrame()
    es_moufid2 = pd.DataFrame()
    #es_moufid2 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Moufid\Repa EvSc Moufid 2aep (Edited _5).csv")
    #es_moufid2["Sèance"] = [json.loads(es_moufid2["Sèance"][i]) for i in range(len(es_moufid2["Sèance"]))]

    es_moufid3 = pd.DataFrame()
    es_moufid4 = pd.DataFrame()
    es_moufid5 = pd.DataFrame()
    es_moufid6 = pd.DataFrame()

    es_moufid_manuel = [es_moufid1,
                        es_moufid2,
                        es_moufid3,
                        es_moufid4,
                        es_moufid5,
                        es_moufid6]

    # Mounir
    es_mounir1 = pd.DataFrame()
    es_mounir2 = pd.DataFrame()
    es_mounir3 = pd.DataFrame()
    es_mounir4 = pd.DataFrame()
    es_mounir5 = pd.DataFrame()
    #es_mounir5 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Mounir\Repa EvSc Mounir 5aep (Edited _6).csv")
    #es_mounir5["Sèance"] = [json.loads(es_mounir5["Sèance"][i]) for i in range(len(es_mounir5["Sèance"]))]

    es_mounir6 = pd.DataFrame()
    es_mounir_manuel = [es_mounir1,
                        es_mounir2,
                        es_mounir3,
                        es_mounir4,
                        es_mounir5,
                        es_mounir6]

    # Marji3:
    es_marjii1 = pd.DataFrame()
    es_marjii2 = pd.DataFrame()
    es_marjii3 = pd.DataFrame()
    #es_marjii4 = pd.read_csv(r"C:\Users\hp\Downloads\edited repartitions\EvSc repa\Marji3\Eveil Scientifique 4 aep ALMARJI3.csv")
    es_marjii4=pd.read_csv("https://docs.google.com/spreadsheets/d/1JCDJDylqfHJp5YlKo8ifrLl1pl12nxDAEoS9hpzosoE"
                           "/gviz/tq?tqx=out:csv&sheet==Eveil_Scientifique_4_aep_ALMARJI3")
    es_marjii4["Sèance"] = [json.loads(es_marjii4["Sèance"][i]) for i in range(len(es_marjii4["Sèance"]))]
    es_marjii5 = pd.DataFrame()
    es_marjii6 = pd.DataFrame()

    es_marjii_manuel = [es_marjii1,
                        es_marjii2,
                        es_marjii3,
                        es_marjii4,
                        es_marjii4,
                        es_marjii5,
                        es_marjii6
                        ]

    EvSc_manuels = {"الجديد": es_jadid_manuel,
                    "المنهل": manhal_manuel,
                    "الفضاء": es_fadaa_manuel,
                    "المفيد": es_moufid_manuel,
                    "المنير": es_mounir_manuel,
                    "المرجع": es_marjii_manuel}

    try:
        if EvSc_manuels[str(manuel)][int(level) - 1].empty:
            return "Ce manuel n'est pas disponible pour ce niveau, choisissez un autre manuel!"

        else:

            return EvSc_manuels[str(manuel)][int(level) - 1]
    except:
        print("Ce manuel n'est pas disponible pour ce niveau, choisissez un autre manuel!")
        print("Remarque: le niveau doit être entre 1 et 6!")








# 2 weeks for ED, 5 weeks for each Unit and 1 week for each EV ,pedagical_week=6 days

def U_W_D(fr_manuel1, fr_manuel2, math_manuel1, math_manuel2, es_manuel1, es_manuel2, C1, C2, séance_de_lundi, u):#,in_t,ou_t,rec_t,switch_t):
    global buffer
    buffer = BytesIO()
    #emplois=pd.read_csv(r"C:\Users\hp\Downloads\emplois_3_4_(2).csv")
    emplois=st.session_state["emplois"]
    emplois.dropna(inplace=True)
    emplois.index = range(emplois.shape[0])
    emplois[["Sèance", "Durèe"]] = emplois[["Sèance", "Durèe"]].astype("int")
    emplois[["Matière", "Sèance", "Durèe"]] = emplois[["Matière", "Sèance", "Durèe"]].astype("str")

    #"""--------------------------------Reparations-------------------------------"""
    #Fr
    my_fr_class_1 = fr_manuel_level(manuel=fr_manuel1,level=C1)
    my_fr_class_2 = fr_manuel_level(manuel=fr_manuel2,level=C2)
    fr_class = [my_fr_class_1, my_fr_class_2]


    #Maths
    my_math_class_1 = maths_manuel_level(manuel=math_manuel1, level=C1)
    my_math_class_2 = maths_manuel_level(manuel=math_manuel2, level=C2)
    math_class = [my_math_class_1, my_math_class_2]

    #Eveil Scientifique
    my_EvSc_class_1 = EvSc_manuel_level(level=C1, manuel=es_manuel1)
    my_EvSc_class_2 = EvSc_manuel_level(level=C2, manuel=es_manuel2)
    EvSc_class = [my_EvSc_class_1, my_EvSc_class_2]

    ##"Eveil scientifique" == "Eveil Scientifique"

    # """------------------Dates-----------------"""
    all_time = np.arange(dt.date(2023, 9, 4), dt.date(2024, 6, 7)).astype(datetime)
    holiday_list = pd.Series({
        "Mawlid_nabawi": [dt.date(2023, 9, 28), dt.date(2023, 9, 29)],
        "V1S1": np.arange(dt.date(2023, 10, 16), dt.date(2023, 10, 22)),
        "March_vert": [dt.date(2023, 11, 6)],
        "Aid_isstiklal": [dt.date(2023, 11, 18)],
        "V2S1": np.arange(dt.date(2023, 12, 4), dt.date(2023, 12, 10)),
        "Bonne_annee": [dt.date(2024, 1, 1)],
        "Watika_Isstiklala": [dt.date(2024, 1, 11)],
        "Amazigh_Day": [dt.date(2024, 1, 14)],
        "Middle_V": np.arange(dt.date(2024, 1, 22), dt.date(2023, 1, 29)),
        "V1S2": np.arange(dt.date(2024, 3, 11), dt.date(2024, 3, 17)),
        "Aid_Fiter": np.arange(dt.date(2024, 4, 7), dt.date(2024, 4, 12)),
        "V2S2": np.arange(dt.date(2024, 4, 29), dt.date(2024, 5, 5)),
        "Aid_Adha": [dt.date(2024, 6, 14), dt.date(2024, 6, 15), dt.date(2024, 6, 17)]
    })

    for i in range(len(holiday_list.values)):
        for j in range(len(holiday_list.values[i])):
            if holiday_list.values[i][j] in all_time:
                all_time = all_time[all_time != holiday_list.values[i][j]]

    No_wekends_all_dates = []
    for date in all_time:

        str_date = date.strftime("%A, %d %B, %Y")
        if ("Sunday" in str_date):
            day = "Weekend!"
        else:
            No_wekends_all_dates.append(date)

    # """------Units, Classes And Manuels------"""

    ED = No_wekends_all_dates[0:12]
    U1 = No_wekends_all_dates[12:42]
    U2 = No_wekends_all_dates[42:72]
    U3 = No_wekends_all_dates[72:102]

    EV_S1 = No_wekends_all_dates[102:108]

    U4 = No_wekends_all_dates[108:138]
    U5 = No_wekends_all_dates[138:168]
    U6 = No_wekends_all_dates[168:198]

    EV_S2 = No_wekends_all_dates[198:204]

    Unites = {"U1": U1, "U2": U2, "U3": U3, "U4": U4, "U5": U5, "U6": U6}
    unite = Unites[u]
    classes = [C1, C2]
    fr_manuels = [fr_manuel1, fr_manuel2]
    math_manuels = [math_manuel1, math_manuel2]
    es_manuels = [es_manuel1, es_manuel2]

    #"""------Sheet Format preparation------"""
    align = Alignment(horizontal="center",
                      vertical="center",
                      wrapText=True)

    thin = Side(border_style="thin", color="4617F1")

    bord = Border(left=thin,
                  right=thin,
                  top=thin,
                  bottom=thin)
    font = Font(name="Lucida Handwriting",sz=12)

    weeks_days = []
    sheets = []
    wb = Workbook()
    for k in range(len(unite)):
        sheets.append(wb.create_sheet(f'sheet {k}'))

    for i, date, sheet in zip(enumerate(unite), unite, sheets):
        date_list = format_datetime(date, 'full', locale='fr_FR').split()[0:4]
        str_date = " ".join(date_list).replace(",", " ")

        for cell in sheet["B4:H7"] + sheet["B8:G8"] :#+ sheet["F2:F3"]:
            for x in cell:
                x.alignment = align
                x.border = bord
        for cell in sheet["B9:F9"]:
            for y in cell:
                y.border = bord
                y.alignment = Alignment(horizontal='left', vertical='top')

        for cell in sheet["B4:H4"] + sheet["B8:H8"]:
            for w in cell:
                w.fill = PatternFill("solid", "FFA500")
        sheet["B6"].fill = PatternFill("solid", "FFA500")

        for cell in sheet["B5:G5"] + sheet["B7:G7"]:
            for v in cell:
                v.font = font

        sheet.column_dimensions["F"].width = 50
        sheet.column_dimensions["D"].width = 22
        sheet.column_dimensions["B"].width = 12
        for lettre in ["C", "E", "G", "H"]:
            sheet.column_dimensions[lettre].width = 8
        sheet.row_dimensions[5].height = 130
        sheet.row_dimensions[7].height = 130

        for row in [5, 7, 9]:
            sheet.row_dimensions[row].height = 100

        # citation
        sheet.merge_cells('B9:E9')
        sheet["B9"] = "CITATION: "

        # remarque
        # sheet.merge_cells('B9:E9')
        sheet["F9"] = "REMARQUES: "

        sheet.merge_cells('B6:H6')
        sheet["B6"] = "RECREATION"

        sheet.merge_cells('G8:H8')
        sheet["G8"] = f'Unité: {1 + list(Unites.values()).index(unite)}'

        sheet["F2"] = f'Date:'#{str_date}'

        #----------------in and out time---------------------

        sheet["B4"] = "Temps: "
        outmor_torec_t = inmor_t + ((outmor_t - inmor_t) - rec_t) / 2
        inmor_fromrec_t = outmor_torec_t + rec_t
        outeven_torec_t = ineven_t + ((outeven_t - ineven_t) - rec_t) / 2
        ineven_fromrec_t = outeven_torec_t + rec_t
        if séance_de_lundi == "Matinèe":

            if date_list[0] in ["lundi", "mercredi", "samedi"]:

                sheet["B5"] = f'De {inmor_t}  à {outmor_torec_t}'
                sheet["B7"] = f'De {inmor_fromrec_t}  à {outmor_t}'

            else:

                sheet["B5"] = f'De {ineven_t}  à {outeven_torec_t}'
                sheet["B7"] = f'De {ineven_fromrec_t}  à {outeven_t}'
        else:

            if date_list[0] in ["mardi", "jeudi", "vendredi"]:
                sheet["B5"] = f'De {inmor_t}  à {outmor_torec_t}'
                sheet["B7"] = f'De {inmor_fromrec_t}  à {outmor_t}'

            else:
                sheet["B5"] = f'De {ineven_t}  à {outeven_torec_t}'
                sheet["B7"] = f'De {ineven_fromrec_t}  à {outeven_t}'




                # Discipline
        sheet["D4"] = "Discipline"

        # Classe
        sheet["C4"] = "Classe"

        # Maths
        def Maths_peda_class(niveau):
            my_class = math_class[niveau]
            if type(my_class) == pd.core.frame.DataFrame:
                maths_Unite_Repa = my_class[my_class["U"] == 1 + list(Unites.values()).index(unite)]
                maths_Week_Repa = maths_Unite_Repa[maths_Unite_Repa["S"] == 1 + int(i[0] / 6)]
                return (maths_Week_Repa)
            else:
                return my_class

        # Frensh
        def fr_peda_class(niveau):
            my_class =fr_class[niveau]
            # my_class.drop(columns=["Unnamed: 0"],inplace=True)
            my_class.index = my_class["S"]
            if type(my_class) == pd.core.frame.DataFrame:
                fr_Unite_Repa = my_class[my_class["U"] == 1 + list(Unites.values()).index(unite)]
                fr_Week_Repa = fr_Unite_Repa[fr_Unite_Repa["S"] == 1 + int(i[0] / 6)]
                return (fr_Week_Repa)
            else:
                return my_class

        # eveil scientifique:

        def EvSc_peda_class(niveau):
            my_class = EvSc_class[niveau]
            if type(my_class) == pd.core.frame.DataFrame:
                EvSc_Unite_Repa = my_class[my_class["U"] == 1 + list(Unites.values()).index(unite)]
                EvSc_Week_Repa = EvSc_Unite_Repa[EvSc_Unite_Repa["S"] == 1 + int(i[0] / 6)]
                return (EvSc_Week_Repa)
            else:
                return my_class

        def courses(niveau, matiéres_cell):
            global classe_emploi
            classe_emploi = emplois[emplois["Niveau"] == classes[niveau]]
            for peda_day in classe_emploi["Jour"]:

                if peda_day == (1 + len(weeks_days) % 6):
                    global jour_emploi
                    jour_emploi = classe_emploi[classe_emploi["Jour"] == peda_day]
                    sheet[matiéres_cell] = '\n'.join(jour_emploi["Matière"].to_list())

        def course_element(niveau, element_cell):
            courses_elements = []
            jour_emploi = classe_emploi[classe_emploi["Jour"] == 1 + len(weeks_days) % 6]
            emplois_week_Maths = classe_emploi[classe_emploi["Matière"] == "Maths"]
            emplois_week_EvSc = classe_emploi[classe_emploi["Matière"] == "Eveil Scientifique"]

            for course in jour_emploi["Matière"]:

                # FR
                fr_repa = fr_peda_class(niveau)
                if course in fr_repa.columns:
                    courses_elements.append(fr_repa[course][1 + int(i[0] / 6)])

                # Maths
                math_repa = Maths_peda_class(niveau)
                if course in math_repa.columns:

                    emplois_jour_Maths = emplois_week_Maths[emplois_week_Maths["Jour"] == 1 + len(weeks_days) % 6]
                    #print(int(list(emplois_jour_Maths["Sèance"])[0]))
                    seance = 6 if emplois_jour_Maths.empty else int(list(emplois_jour_Maths["Sèance"])[0])
                    seance = str(seance)

                    for s in list(math_repa["Sèance"]):

                        if seance in s:
                            indices = math_repa[math_repa['Sèance'] \
                                .apply(lambda x: any(item in x for item in s))].index
                            courses_elements.append(math_repa["Maths"][indices[0]])

                # Eveil Scientifique:
                es_repa = EvSc_peda_class(niveau)
                if course in es_repa.columns:
                    emplois_jour_EvSc = emplois_week_EvSc[emplois_week_EvSc["Jour"] == 1 + len(weeks_days) % 6]
                    seance = 6 if emplois_jour_EvSc.empty else int(emplois_jour_EvSc["Sèance"])
                    for s in list(es_repa["Sèance"]):
                        if seance in s:
                            indices = es_repa[es_repa['Sèance'].apply(lambda x: any(item in x for item in s))].index
                            courses_elements.append(es_repa["Eveil Scientifique"][indices[0]])
            title = fr_peda_class(niveau)["U. D / Thème"][1 + int(i[0] / 6)]

            sheet["F3"] = f'Théme:  {title}'

            course_element = [str(cours_elem) for cours_elem in courses_elements]
            sheet[element_cell] = '\n'.join(course_element)
            #print('\n'.join(course_element))

        def periode_seance(niveau, period_cell, seance_cell):
            for peda_day in classe_emploi["Jour"]:
                if peda_day == (1 + len(weeks_days) % 6):
                    global jour_emploi
                    jour_emploi = classe_emploi[classe_emploi["Jour"] == peda_day]

                    sheet[period_cell] = '"\n'.join(jour_emploi["Durèe"].to_list())
                    sheet[seance_cell] = '\n'.join(jour_emploi["Sèance"].to_list())

        if date_list[0] in ["lundi", "mercredi", "vendredi"]:

            sheet["C5"] = classes[0]
            # try:
            courses(0, "D5")  # ,"F5")
            periode_seance(0, "E5", "G5")
            course_element(0, "F5")

            sheet["C7"] = classes[1]
            courses(1, "D7")
            periode_seance(1, "E7", "G7")
            course_element(1, "F7")
            # except:
            # print("error!")
            # break

        else:
            # try:
            sheet["C5"] = classes[1]
            courses(1, "D5")
            periode_seance(1, "E5", "G5")
            course_element(1, "F5")

            sheet["C7"] = classes[0]
            courses(0, "D7")
            periode_seance(0, "E7", "G7")
            course_element(0, "F7")
            # except:
            # print("error!")
            # break

        # Théme
        # title = peda_class(classes[niveau])["U. D / Thème"][1]
        # print("Théme:",title)
        # sheet["F3"] = f'Théme:  {title}'

        # Durée
        sheet["E4"] = "Durée"

        # Elements du cours
        sheet["F4"] = "Elements du cours (Intitulé / Objectifs / Etapes..)"

        # Nombre de Séance
        sheet["G4"] = "Séance"

        # Nombre de fiche
        sheet["H4"] = "Fiche"

        # Smaine
        sheet["F8"] = f'Semaine:{1 + int(i[0] / 6)}'
        # print("Day: ", 1 + len(weeks_days) % 6)

        # Jour


        sheet.merge_cells('B8:E8')
        sheet["B8"] = f'Jour: {1 + len(weeks_days) % 6}'
        weeks_days.append(i[0])

        #print("----------------------------")

    # wb.save(f'Cahier_Journalier_U{1+list(Unites.values()).index(unite)} niveaux {C1}-{C2}.xlsx')
    wb.save(buffer)
    return buffer

#j1=U_W_D(fr_manuel1="Mes apprentissages", fr_manuel2 = "Mes apprentissages", math_manuel1="الفضاء", math_manuel2="الجيد", es_manuel1="المنهل", es_manuel2="المرجع", C1=3, C2=4, séance_de_lundi="Matinée", u="U2")
#print(j1)

# """"----------------Streamlit App-----------------"""

# Title

st.title("Le Cahier Journalier Du Professeur De Cycle Primaire ")
st.empty()

dic = {"Niveau": [3, 4], "Français": ["Mes apprentissages", "Mes apprentissages"],
       "Maths": ["الفضاء", "الجيد"], "Eveil Scientifique": ["المنهل", "المرجع"]}
dispo_manuels_df = pd.DataFrame(dic, index=None)
dispo_manuels = np.array(list(dic.values()))
dispo_manuels = dispo_manuels.reshape(1, 4 * dispo_manuels.shape[1])[0]
#print("dispo: ", dispo_manuels)

# Select Teaching Language
col1,col2=st.columns([0.7, 0.3])
form = col1.form(key="my_form")


teaching_language = form.selectbox("Sèlèctionez la langue ", ["Français", "Arabe"], placeholder="choisissez une option",
                                   index=None)

next = form.form_submit_button("Cliquez pour choisir les niveaux ou les groupes")
# level1 = None
# level2 = None
# unit = None
# period = None
# submit = None

# Define function to select options
# Select Manual Names
if (teaching_language == "Français"):
    # Select Levels
    level1 = form.multiselect(
        "Choisissez un niveau ou un groupe :", [3, 4], max_selections=1)  # , default=3)
    # [1,2,5,6]
    level2 = form.multiselect(
        "Choisissez un niveau ou un groupe:", [3, 4], max_selections=1)  # , default=4)  # [1,2,5,6]
    form.form_submit_button(label="Cliquez pour choisir les manuels")

    selected_manuels = []

    if (len(level1) != 0) and (len(level2) != 0):

        # disponible manuels:
        st.write("Les manuels disponibles: ")
        st.table(dispo_manuels_df)
        # st.dataframe(dispo_manuels_df)

        # Maths
        math_manuels = ["المفيد", "الجيد", "الفضاء"]
        maths_manual_name1 = form.multiselect(f"Choisissez le manuel de maths pour le niveau: {level1[0]}",
                                              math_manuels,
                                              key="math1", max_selections=1)

        maths_manual_name2 = form.multiselect(f"Choisissez le manuel de maths pour le niveau: {level2[0]}",
                                              math_manuels,
                                              key="math2", max_selections=1)

        # French
        french_manuels = [
            "Mes apprentissages"]  # , "L'oasis des mots","Pour communiquer", "L'école des mots", "Le nouvel espace","Mon livre de français"]
        french_manual_name1 = form.multiselect(f"Choisissez le manuel de français pour le niveau:{level1[0]}",
                                               french_manuels, key="fr1", max_selections=1)
        french_manual_name2 = form.multiselect(f"Choisissez le manuel de français pour le niveau:{level2[0]}",
                                               french_manuels, key="fr2", max_selections=1)

        # Act. Sc.
        Act_Sc_manuels = ["الجديد", "المنهل", "الفضاء", "المفيد", "المنير", "المرجع"]
        act_sc_manual_name1 = form.multiselect(f"Choisissez le manuel d'éveil scientifique pour le niveau:{level1[0]}",
                                               Act_Sc_manuels, key="evsc1", max_selections=1)

        act_sc_manual_name2 = form.multiselect(f"Choisissez le manuel d'éveil scientifique pour le niveau:{level2[0]}",
                                               Act_Sc_manuels, key="evsc2", max_selections=1)
        # Select Period
        period = form.selectbox("sélectionnez la sèance", ["Matinée", "Après midi"])
        # Select Unit
        unit = form.multiselect("sélectionnez l'unitè", ["U1", "U2", "U3", "U4", "U5", "U6"],
                                max_selections=1)  # , default=["U1"])
        # select enter time on the morning
        inmor_t = form.time_input("Sélectionnez l'heure d'entrée le matin", dt.time(hour=9, minute=0),
                                key="inmor_t",step=dt.timedelta(minutes=5))
        inmor_t = dt.timedelta(hours=inmor_t.hour, minutes=inmor_t.minute)
        # select the out time on the morning
        outmor_t = form.time_input("sélectionnez l'heure de sortie le matin", dt.time(hour=13, minute=30),
                                 key="outmor_t",step=dt.timedelta(minutes=5))
        outmor_t = dt.timedelta(hours=outmor_t.hour,minutes=outmor_t.minute)
        # select the in time on the evening
        ineven_t = form.time_input("Sélectionnez l'heure d'entrée le soir", dt.time(hour=13, minute=40),
                                 key="ineven_t",step=dt.timedelta(minutes=5))
        ineven_t = dt.timedelta(hours=ineven_t.hour, minutes=ineven_t.minute)
        # select the out time on the evening
        outeven_t =  form.time_input("Sélectionnez l'heure de sortie le soir", dt.time(hour=18, minute=30),
                                   key="outeven_t",step=dt.timedelta(minutes=5))
        outeven_t = dt.timedelta(hours=outeven_t.hour, minutes=outeven_t.minute)
        # select the recreation time
        rec_t = form.time_input("Sélectionnez le temps de la rècreation:",dt.time(minute=10),
                                key= "rec_t",step=dt.timedelta(minutes=1))
        rec_t = dt.timedelta(minutes=rec_t.minute)

        enregistre = form.form_submit_button("Enregistrez!")

        options = [french_manual_name1, french_manual_name2,
                   maths_manual_name1, maths_manual_name2,
                   act_sc_manual_name1, act_sc_manual_name2]
        if all(len(val) != 0 for val in options) and (enregistre == True):

            selected_manuels.extend([level1, level2, french_manual_name1, french_manual_name2,
                                     maths_manual_name1, maths_manual_name2, act_sc_manual_name1,
                                     act_sc_manual_name2])

            if all(str(value[0]) in dispo_manuels for value in selected_manuels):
                # """"-------------Emplois Du Temps-----------""""
                if "emplois" in st.session_state:
                    emplois = st.session_state["emplois"]
                    emplois.dropna(inplace=True)
                    emplois.index = range(emplois.shape[0])
                    emplois[["Sèance", "Durèe"]] = emplois[["Sèance", "Durèe"]].astype("int")
                    emplois[["Matière", "Sèance", "Durèe"]] = emplois[["Matière", "Sèance", "Durèe"]].astype("str")
                    with st.spinner("La création de votre journal est en cours..."):
                        journal = U_W_D(
                          fr_manuel1=french_manual_name1[0],
                          fr_manuel2=french_manual_name2[0],
                          math_manuel1=maths_manual_name1[0],
                          math_manuel2=maths_manual_name2[0],
                          es_manuel1=act_sc_manual_name1[0],
                          es_manuel2=act_sc_manual_name2[0],
                          C1=level1[0],
                          C2=level2[0],
                          séance_de_lundi=period,
                          u=unit[0],
                          )
                    st.success("Votre journal est a été crèer avec succès! ")
                    download = st.download_button(
                       label="Téléchargez votre cahier de journal",
                       data=journal,
                       key="workbook.xlsx",
                       file_name=f"Cahier_Journalier_{unit[0]}_niveaux {level1[0]}-{level2[0]}_{period}.xlsx",
                        )
                else:
                    st.warning("Veuillez importez ou créer votre emplois d'abord!")
            else:
                form.warning("Le(s) manuel(s) sélectionné(s) est\sont indisponible(s) pour "
                             "le moment. Veuillez choisir de(s) autre(s)! ")
        else:
            form.warning("Veuillez sélectionner toutes les options!")

    else:
        form.warning("Sélectionnez les niveaux!")

elif (teaching_language == "Arabe"):
    form.warning("La création du cahier de journal de la langue arabe est impossible pour le moment!")

col2.markdown("""
<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-4265574502229447"
    crossorigin="anonymous"></script>""",unsafe_allow_html=True
             )
